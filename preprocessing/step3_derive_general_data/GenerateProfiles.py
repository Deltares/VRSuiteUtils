import copy
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
import pandas as pd
import re
from pathlib import Path
import csv

"""This is a routine to extract profiles from AHN data. It is based on detecting points with the highest covariance in 
slope, which are the crest, toe and berm lines. For some profiles manual point adjustments have been made. Therefore: 
handle this with care and check visually whether generated profiles are correct!!!"""

def correct_raw_data(elevation_data):
    # round x coordinates to integers (why?)
    profile_data = copy.deepcopy(elevation_data)
    profile_data["x"] = profile_data["x"].round(0).astype(int)
    profile = pd.DataFrame(
        {
            "x": [
                i
                for i in range(profile_data["x"].iloc[0], profile_data["x"].iloc[-1], 2)
            ]
        }
    )
    profile = pd.merge(profile, profile_data, how="left", on="x").interpolate()
    return profile

def compute_statistics(profile_data, window=2):
    # Smoothen of the profile and calculate its derivative and covariation
    profile_data["z_averaged"] = profile_data["z"].rolling(window=window, min_periods=1, center=True).mean()
    profile_data["z_derivative"] = profile_data["z_averaged"].diff()
    profile_data["z_covariance"] = (
        profile_data["z"].rolling(window=window * 2, min_periods=1, center=True).cov()
    )
    return profile_data


def identify_slopes(profile_data):
    # Identify the slopes from the covariance plot THIS IS THE OLD VERSION!
    slopes = (
        profile_data[(profile_data["z_covariance"] >= 0.15) & (profile_data["x"] >= -50) & (profile_data["x"] <= 75)]
        .reset_index()
        .rename(columns={"index": "profile index"})
    )
    #FROM HERE IT IS NEW:
    #get outer_slope from profile_data. It must have x=0 at upper bound and lower bound is when z_covariance < 0.1
    outer_slope = profile_data.loc[(profile_data["x"] <= 0) & (profile_data["z_covariance"] > 0.1)]
    #drop all points from outer_slope that do not have a steadily increasing dataframe index
    index_diff = np.append(np.diff(outer_slope.index),1)
    outer_slope = outer_slope.loc[index_diff==1]
    BUK = outer_slope.iloc[-1]
    BUT = outer_slope.iloc[0]

    #find inner crest, which should be as high or lower than outer crest
    BIK = profile_data.loc[(profile_data["x"] > 1) & (profile_data["z_averaged"] < outer_slope.z.values[-1])].iloc[0]

    #for inner_slope, find first point with x>0 where z_covariance > 0.05
    inner_slope = profile_data.loc[(profile_data["x"] >= BIK.x) & (profile_data["z_covariance"] > 0.025)]

    #check if x at inner slope has gaps larger than 5 meters, this indicates a berm:
    index_diff = np.append(np.diff(inner_slope.index),1)
    berm_start = inner_slope.loc[index_diff>5] #this is the BBL
    if not berm_start.empty:
        berm_found = True
        #berm found so add BBL and EBL
        BBL = berm_start.iloc[0]
        EBL = inner_slope.loc[inner_slope["x"] > BBL.x].iloc[0]
    else:
        berm_found = False
    BIT = inner_slope.iloc[-1]

    #for testing:
    plt.plot(profile_data.x, profile_data.z)
    plt.plot(BUT.x, BUT.z, 'or')
    plt.plot(BUK.x, BUK.z, 'or')
    plt.plot(BIK.x, BIK.z, 'or')
    try:
        plt.plot(BBL.x, BBL.z, 'or')
        plt.plot(EBL.x, EBL.z, 'or')
    except:
        pass
    plt.plot(BIT.x, BIT.z, 'or')

    profile_points = {'BUT': BUT, 'BUK': BUK, 'BIK': BIK, 'BIT': BIT}
    if berm_found:
        profile_points['BBL'] = BBL
        profile_points['EBL'] = EBL
    return profile_points

def plot_profile(elevation_data, profile_points, title, path):
    fig,ax = plt.subplots()
    ax.plot(elevation_data.x, elevation_data.z)
    for point in profile_points.keys():
        ax.plot(profile_points[point].x, profile_points[point].z, 'or')
    ax.set_xlim(left = profile_points['BUT'].x-10,right = profile_points['BIT'].x+10)
    plt.savefig(path.joinpath(title + '.png'))
    plt.close()
def extract_profile_data(elevation_data, window, title, path):
    #correct the elevation data
    profile = correct_raw_data(elevation_data)

    #compute the statistics of the elevation
    profile = compute_statistics(profile,window=window)

    profile_points = identify_slopes(profile)

    plot_profile(elevation_data, profile_points, title, path)
    return profile_points


    #in principe everything below here is not needed anymore, if the approach above is robust enough
    
    # Correct the selected slopes on gradient direction to get rid of non-levee slopes
    slopes = slopes[
        ((slopes["x"] < 0) & (slopes["z_derivative"] > 0))
        | ((slopes["x"] > 0) & (slopes["z_derivative"] < 0))
    ].reset_index(drop=True)
    pivot_index = (
        slopes[slopes["profile index"].diff() > 1]
        .reset_index()
        .rename(columns={"index": "slope index"})
    )

    # Correct selected slopes on covariance intensity to get rid of small disturbances like ditches
    slopes_corrected = copy.deepcopy(slopes)
    for i in range(len(pivot_index)):
        if i == 0:
            first_slope_intensity = slopes.iloc[
                0 : int(pivot_index.iloc[i]["slope index"])
            ]["z_covariance"].max()
            if first_slope_intensity < 0.3:
                slopes_corrected = slopes_corrected.drop(
                    slopes.iloc[0 : int(pivot_index.iloc[i]["slope index"])].index
                )

        elif i != 0 and i != len(pivot_index) - 1:
            intermediate_slope_intensity = slopes.iloc[
                int(pivot_index.iloc[i - 1]["slope index"]) : int(
                    pivot_index.iloc[i]["slope index"]
                )
            ]["z_covariance"].max()
            if intermediate_slope_intensity < 0.3:
                slopes_corrected = slopes_corrected.drop(
                    slopes.iloc[
                        int(pivot_index.iloc[i - 1]["slope index"]) : int(
                            pivot_index.iloc[i]["slope index"]
                        )
                    ].index
                )

        elif i != 0 and i == len(pivot_index) - 1:
            intermediate_slope_intensity = slopes.iloc[
                int(pivot_index.iloc[i - 1]["slope index"]) : int(
                    pivot_index.iloc[i]["slope index"]
                )
            ]["z_covariance"].max()
            if intermediate_slope_intensity < 0.3:
                print(slopes_corrected)
                slopes_corrected = slopes_corrected.drop(
                    slopes.iloc[
                        int(pivot_index.iloc[i - 1]["slope index"]) : int(
                            pivot_index.iloc[i]["slope index"]
                        )
                    ].index
                )

            last_slope_intensity = slopes.iloc[
                int(pivot_index.iloc[len(pivot_index) - 1]["slope index"] - 1) :
            ]["z_covariance"].max()
            if last_slope_intensity < 0.3:
                print(slopes_corrected)
                slopes_corrected = slopes_corrected.drop(
                    slopes.iloc[
                        int(pivot_index.iloc[len(pivot_index) - 1]["slope index"] - 1) :
                    ].index
                )

    # Overwrite with the corrected versions and calculate the pivot indexes of the profile
    slopes = slopes_corrected.reset_index(drop=True)
    if len(slopes > 0):
        pivot_indices = [int(slopes["profile index"][0])]
        for i in range(1, len(slopes) - 1):
            if (slopes["profile index"][i + 1] - slopes["profile index"][i - 1]) != 2:
                pivot_indices.append(int(slopes["profile index"][i]))
        pivot_indices.append(int(slopes["profile index"].iloc[-1]))
    # the pivot index is the index of the slope that is the first slope of a new profile
    pivot_index = (
        slopes[slopes["profile index"].diff() > 1]
        .reset_index()
        .rename(columns={"index": "slope index"})
    )

    # Find the profile coordinates of the levee with respect to the outer_crest_index
    outer_crest_index = (
        (df["x"] - 0).abs().idxmin()
    )  # Find the index of where x is closest to 0
    matched_pivot_index = int(
        (pivot_index["profile index"] - outer_crest_index).abs().idxmin()
    )
    matched_slope_index = int(pivot_index.iloc[matched_pivot_index]["slope index"])
    print(pivot_indices)
    # In case of no berm, len(pivot indices) = 4
    if len(pivot_indices) == 4:
        outer_toe_index = int(pivot_indices[0])
        # outer_crest_index = int(slopes.iloc[pivot_indices[1]]['profile index'])
        inner_crest_index = int(pivot_indices[2])
        inner_toe_index = int(pivot_indices[3])
    # to do: change inner and outer to water and land
    # In case of a berm, find if berm is outer or inner
    elif len(pivot_indices) == 6:
        if sum(i < outer_crest_index for i in pivot_indices) > 3:
            # outer berm present
            print("outer berm present")
            outer_toe_index = int(pivot_indices[0])
            berm_outer_toe_side_index = int(pivot_indices[1])
            berm_outer_crest_side_index = int(pivot_indices[2])
            inner_crest_index = int(pivot_indices[4])
            inner_toe_index = int(pivot_indices[5])
        elif sum(i < outer_crest_index for i in pivot_indices) < 3:
            # inner berm present
            print("inner berm present")
            outer_toe_index = int(pivot_indices[0])
            inner_crest_index = int(pivot_indices[2])
            berm_inner_crest_side_index = int(pivot_indices[3])
            berm_inner_toe_side_index = int(pivot_indices[4])
            inner_toe_index = int(pivot_indices[5])

    # Manual corrections
    cross_section = int(re.findall("\d+", title)[0])
    # if cross_section == 1: inner_toe_index = 85
    # if cross_section == 6: inner_toe_index = 83
    # if cross_section == 15: inner_toe_index = 83

    # Extract x and z coordinates of the profile coordinates
    if len(pivot_indices) == 4:
        x_coords = [
            df.loc[outer_toe_index]["x"],
            df.loc[outer_crest_index]["x"],
            df.loc[inner_crest_index]["x"],
            df.loc[inner_toe_index]["x"],
        ]
        z_coords = [
            df.loc[outer_toe_index]["z"],
            max(df.loc[outer_crest_index]["z"], df.loc[inner_crest_index]["z"]),
            max(df.loc[outer_crest_index]["z"], df.loc[inner_crest_index]["z"]),
            df.loc[inner_toe_index]["z"],
        ]
        # x_coords = [df.loc[inner_toe_index]['x'], df.loc[inner_crest_index]['x'], df.loc[outer_crest_index]['x'], df.loc[outer_toe_index]['x']]
        # z_coords = [df.loc[inner_toe_index]['z'], (df.loc[inner_crest_index]['z'] + df.loc[outer_crest_index]['z']) / 2, (df.loc[inner_crest_index]['z'] + df.loc[outer_crest_index]['z']) / 2, df.loc[outer_toe_index]['z']]
    elif (len(pivot_indices) == 6) and (
        sum(i < outer_crest_index for i in pivot_indices) < 3
    ):
        x_coords = [
            df.loc[outer_toe_index]["x"],
            df.loc[outer_crest_index]["x"],
            df.loc[inner_crest_index]["x"],
            df.loc[berm_inner_crest_side_index]["x"],
            df.loc[berm_inner_toe_side_index]["x"],
            df.loc[inner_toe_index]["x"],
        ]

        z_coords = [
            df.loc[outer_toe_index]["z"],
            max(df.loc[outer_crest_index]["z"], df.loc[inner_crest_index]["z"]),
            max(df.loc[outer_crest_index]["z"], df.loc[inner_crest_index]["z"]),
            (
                df.loc[berm_inner_crest_side_index]["z"]
                + df.loc[berm_inner_toe_side_index]["z"]
            )
            / 2,
            (
                df.loc[berm_inner_crest_side_index]["z"]
                + df.loc[berm_inner_toe_side_index]["z"]
            )
            / 2,
            df.loc[inner_toe_index]["z"],
        ]

    elif (len(pivot_indices) == 6) and (
        sum(i < outer_crest_index for i in pivot_indices) > 3
    ):
        x_coords = [
            df.loc[outer_toe_index]["x"],
            df.loc[berm_outer_toe_side_index]["x"],
            df.loc[berm_outer_crest_side_index]["x"],
            df.loc[outer_crest_index]["x"],
            df.loc[inner_crest_index]["x"],
            df.loc[inner_toe_index]["x"],
        ]
        z_coords = [
            df.loc[outer_toe_index]["z"],
            (
                df.loc[berm_outer_toe_side_index]["z"]
                + df.loc[berm_outer_crest_side_index]["z"]
            )
            / 2,
            (
                df.loc[berm_outer_toe_side_index]["z"]
                + df.loc[berm_outer_crest_side_index]["z"]
            )
            / 2,
            max(df.loc[outer_crest_index]["z"], df.loc[inner_crest_index]["z"]),
            max(df.loc[outer_crest_index]["z"], df.loc[inner_crest_index]["z"]),
            df.loc[inner_toe_index]["z"],
        ]

        # z_coords = [df.loc[outer_toe_index]['z'], df.loc[berm_outer_toe_side_index]['z'], df.loc[berm_outer_toe_side_index]['z'], df.loc[outer_crest_index]['z'], df.loc[inner_crest_index]['z'], df.loc[inner_toe_index]['z']]

    # Plot the calculated profile coordinates, connected by a dotted line over the measured profile and the covariance of the profile over the cross-section
    fig = plt.figure(figsize=(14, 6))

    ax1 = fig.add_subplot(2, 1, 1)  # 2 rows, 1 column, subplot 1
    ax2 = fig.add_subplot(2, 1, 2)  # 2 rows, 1 column, subplot 2

    ax1.plot(profile_data["x"], profile_data["z"])
    ax1.plot(df["x"], df["z_averaged"], "r")
    ax1.plot(x_coords, z_coords, color="k", marker="o", linestyle="--")
    ax1.set_ylabel("m NAP")
    ax1.set_xlim(min(df["x"]), max(df["x"]))
    ax1.grid()

    ax2.plot(df["x"], df["z_covariance"])
    ax2.set_ylabel("CoV profiel")
    ax2.set_xlim(min(df["x"]), max(df["x"]))
    ax2.grid()

    fig.suptitle(title)
    plt.savefig(path.joinpath(title + ".png"))
    plt.close()

    return pd.DataFrame({"x": x_coords, "z": z_coords})


def main():
    traject = "38-1"
    input_path = Path(r"c:\Users\klerk_wj\OneDrive - Stichting Deltares\00_Projecten\11_VR_HWBP\test_profielen")
    output_path = input_path.joinpath("profiles")
    input_file_name = Path(r"traject_profiles_point.csv")
    output_filename = "Dijkvakindeling_v5.2.xlsx"  # wat gebeurt hiermee?

    # Make output folder if not exist:
    if not output_path.is_dir():
        output_path.mkdir(parents=True, exist_ok=True)

    with open(input_path.joinpath(input_file_name), "r") as csvfile:
        reader = csv.reader(csvfile, delimiter=",")
        header = next(reader)
        for line in reader:
            x_coords = []
            y_coords = []
            z_coords = []
            x = np.zeros(len(line[7:]))
            x[0] = -float(line[6])
            for i, column in enumerate(line[7:]):
                values = re.split(r"\s+", column.strip("[]"))
                x_coords.append(float(values[0]))
                y_coords.append(float(values[1]))
                z_coords.append(float(values[2]))
                if i > 0:
                    x[i] = x[i - 1] + np.sqrt(
                        (x_coords[i] - x_coords[i - 1]) ** 2
                        + (y_coords[i] - y_coords[i - 1]) ** 2
                    )

            profile_temp = {"x": x, "z": z_coords}
            profile_df = pd.DataFrame(profile_temp)
            profile_number = str(line[0])
            print(profile_number)
            profile = extract_profile_data(profile_df, window=2, title="Dwarsprofiel_" + profile_number,
                                           path=output_path)

            # Save extracted data in .csv
            pd.DataFrame.from_dict(profile).transpose()[['x','z']].to_csv(output_path.joinpath(profile_number + ".csv"))

        exit()
        # Write profile_numbers to output file
        wb = load_workbook(input_path.parent.joinpath((output_filename)))
        ws = wb["Dijkvakindeling_keuze_info"]

        for row in range(ws.max_row):
            if ws[row + 1][0].value == i:
                ws.cell(row=row + 1, column=11).value = profile_number
            else:
                pass

        wb.save(input_path.joinpath((output_filename)))


if __name__ == "__main__":
    main()
